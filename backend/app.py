#!/usr/bin/env python3
"""
Flask backend for GST Invoice PDF generation and Salary Excel management
"""

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from datetime import date
import os
import io
import traceback
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from invoice_generator import generate_invoice_pdf

app = Flask(__name__)
CORS(app)

# Create data directory if it doesn't exist
DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)
SALARY_FILE = DATA_DIR / 'salary.xlsx'

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'}), 200

def init_salary_sheet(wb):
    """Initialize salary sheet with headers"""
    ws = wb.active
    ws.title = 'Salary'

    headers = ['Labour Name', 'Salary', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    return ws

@app.route('/api/salary/save', methods=['POST'])
def save_salary():
    """
    Save salary data to Excel file

    Expected JSON:
    {
        "name": "John Doe",
        "salary": 5000,
        "date": "2025-06-16"
    }
    """
    try:
        data = request.json

        if not data.get('name') or not data.get('salary') or not data.get('date'):
            return jsonify({'error': 'Missing required fields: name, salary, date'}), 400

        # Load or create workbook
        if SALARY_FILE.exists():
            wb = load_workbook(SALARY_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = init_salary_sheet(wb)

        # Add new row
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = data['name']
        ws.cell(row=next_row, column=2).value = float(data['salary'])
        ws.cell(row=next_row, column=3).value = data['date']

        # Format number cells
        ws.cell(row=next_row, column=2).number_format = '#,##0.00'

        # Save file
        wb.save(SALARY_FILE)

        return jsonify({'success': True, 'message': 'Salary record added successfully'}), 200

    except Exception as e:
        print(f"Error saving salary: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/salary/get', methods=['GET'])
def get_salary_data():
    """Retrieve all salary records from Excel"""
    try:
        if not SALARY_FILE.exists():
            return jsonify({'records': []}), 200

        wb = load_workbook(SALARY_FILE)
        ws = wb.active
        records = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value:
                records.append({
                    'name': row[0].value,
                    'salary': row[1].value,
                    'date': row[2].value
                })

        return jsonify({'records': records}), 200

    except Exception as e:
        print(f"Error retrieving salary data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-invoice', methods=['POST'])
def generate_invoice():
    """
    Generate PDF invoice from JSON data
    
    Expected JSON:
    {
        "invoiceNo": "INV-2025-001",
        "date": "2025-06-16",
        "customer": {
            "name": "ABC Corp",
            "address": "...",
            "mobile": "...",
            "gst": "..."
        },
        "items": [
            {
                "name": "...",
                "quantity": 1,
                "rate": 100,
                "gstPercent": 9,
                "amount": 109
            }
        ],
        "totals": {
            "subtotal": 100,
            "cgst": 4.5,
            "sgst": 4.5,
            "grandTotal": 109
        }
    }
    """
    try:
        data = request.json
        
        # Validate required fields
        if not data.get('customer') or not data.get('items'):
            return jsonify({'error': 'Missing customer or items data'}), 400
        
        # Generate PDF
        pdf_buffer = generate_invoice_pdf(data)
        
        # Return PDF
        invoice_no = data.get('invoiceNo', 'invoice').replace('/', '-')
        filename = f"{invoice_no}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error generating invoice: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
